"""Synthetic entity tagger (v4_GB).

GB-specific version that only pushes 6 tags:
  Segment, Bu, critical_classification, Region, SupportGroup, ValueStream

Flow:
  1. Read input CSV (name, Business Application, Value Stream)
  2. Query CMDB API for tag data per Business Application
  3. Query NR entity search to find GUIDs by synthetic name
  4. Query NR for existing tags → save backup
  5. Compare CMDB data vs NR tags → generate comparison + push preview
  6. Push changes when apply_tags=true

Configuration: ./config/config.json
"""

import base64
import csv
import json
import os
import time
from datetime import datetime

import pyodbc
import requests
from openpyxl import Workbook, load_workbook

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_CONFIG_PATH = os.path.join(_SCRIPT_DIR, "config", "config.json")
_INPUT_FILE = os.path.join(_SCRIPT_DIR, "GB Synthetics - ManuConnect.csv")
EXTRACTS_DIR = os.path.join(_SCRIPT_DIR, "extracts")

# Only these 6 tags for GB (+ BusinessApplicationCIName for specific apps)
NR_TAG_NAMES = [
    "Segment",
    "Bu",
    "critical_classification",
    "Region",
    "SupportGroup",
    "ValueStream",
    "BusinessApplicationCIName",
]

# CMDB alias map: if the original name fails, try this alternative
CMDB_ALIASES = {
    "Manuconnect Member Correspondence Enablement - CDNGB": "Manuconnect Suite - CDNGB",
}

# For these business apps, hardcode BusinessApplicationCIName to the original CSV name
HARDCODE_BACI_NAME = {
    "Manuconnect Member Correspondence Enablement - CDNGB": "Manuconnect Member Correspondence Enablement - CDNGB",
}


def load_config():
    with open(_CONFIG_PATH) as f:
        return json.load(f)


def clean_criticality(value):
    """Strip '2-', '3-', '4-' prefix from criticality values."""
    if not value:
        return value
    for prefix in ("2-", "3-", "4-"):
        if value.startswith(prefix):
            return value[len(prefix):]
    return value


def _transform_bu(bu, segment):
    """Transform BU value based on segment."""
    if not bu:
        return bu
    seg_lower = (segment or "").lower()
    if "asia" in seg_lower:
        if "," in bu:
            return bu.split(",")[0].strip()
    elif "canada" in seg_lower:
        if "-" in bu:
            return bu.split("-", 1)[1].strip()
    return bu


# ---------------- Input parsing ----------------

def parse_input_file(filepath):
    """Parse input CSV. Returns list of dicts with name, business_app, value_stream."""
    rows = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append({
                "name": row["name"].strip(),
                "business_app": row["Business Application"].strip(),
                "value_stream": row.get("Value stream", row.get("Value Stream", "")).strip(),
            })
    return rows


# ---------------- CMDB API ----------------

def query_cmdb(base_url, app_name, subscription_key=None):
    """Query CMDB Curated Extract API by app_name. Returns the first matching app dict or None."""
    url = f"{base_url}/apps/"
    params = {"app_name": app_name, "format": "json"}
    headers = {}
    if subscription_key:
        headers["Ocp-Apim-Subscription-Key"] = subscription_key
    try:
        response = requests.get(url, params=params, headers=headers, timeout=30, verify=False)
        response.raise_for_status()
        raw = response.json()
        data = raw.get("data", raw) if isinstance(raw, dict) else raw
        if not data:
            return None
        for app in data:
            if app.get("app_name", "").strip().lower() == app_name.lower():
                return app
        return data[0]
    except Exception as e:
        print(f"  [WARN] CMDB query failed for '{app_name}': {e}")
        return None


def extract_tags_from_cmdb(cmdb_data, value_stream):
    """Extract only the 6 GB tags from CMDB response + value_stream from input."""
    if not cmdb_data:
        return {}

    # Find PROD app service
    prod_service = None
    for svc in cmdb_data.get("app_services", []):
        if svc.get("mfc_env", "").upper() == "PROD":
            prod_service = svc
            break

    tags = {
        "Segment": cmdb_data.get("leanix_segment", cmdb_data.get("segment", "")),
        "Bu": _transform_bu(cmdb_data.get("leanix_bu", cmdb_data.get("bu", "")), cmdb_data.get("leanix_segment", cmdb_data.get("segment", ""))),
        "critical_classification": clean_criticality(cmdb_data.get("application_criticality", "")),
        "Region": "PROD",
        "SupportGroup": prod_service.get("app_service_support_group", "") if prod_service else "",
    }

    # Only add ValueStream if non-empty
    if value_stream:
        tags["ValueStream"] = value_stream

    # Remove empty values (don't push empty tags)
    return {k: v for k, v in tags.items() if v}


# ---------------- NR: Entity search by name ----------------

def search_synthetic_guid(api_key, nerdgraph_url, synthetic_name):
    """Search NR for a synthetic monitor by name. Returns GUID or None."""
    escaped_name = synthetic_name.replace("\\", "\\\\").replace('"', '\\"').replace("'", "\\'")
    query = """
    {
      actor {
        entitySearch(query: "name LIKE '%%%s%%' AND domain = 'SYNTH'") {
          results {
            entities {
              guid
              name
            }
          }
        }
      }
    }
    """ % escaped_name

    for attempt in range(1, 4):
        try:
            response = requests.post(
                nerdgraph_url,
                json={"query": query},
                headers={"Api-Key": api_key, "Content-Type": "application/json"},
                timeout=30,
            )
            result = response.json()
            entities = (
                (result.get("data") or {})
                .get("actor", {})
                .get("entitySearch", {})
                .get("results", {})
                .get("entities", [])
            )
            if not entities:
                return None
            for ent in entities:
                if ent.get("name", "").strip() == synthetic_name:
                    return ent["guid"]
            return entities[0]["guid"]
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout, ConnectionResetError) as e:
            if attempt < 3:
                time.sleep(attempt * 5)
                continue
            raise
    return None


# ---------------- NR: Query existing tags ----------------

def query_entity_tags(api_key, nerdgraph_url, guid):
    """Query NR for all tags on an entity. Returns dict {tag_key: [values]}."""
    query = """
    {
      actor {
        entity(guid: "%s") {
          tags {
            key
            values
          }
        }
      }
    }
    """ % guid

    for attempt in range(1, 4):
        try:
            response = requests.post(
                nerdgraph_url,
                json={"query": query},
                headers={"Api-Key": api_key, "Content-Type": "application/json"},
                timeout=30,
            )
            result = response.json()
            entity = (result.get("data") or {}).get("actor", {}).get("entity")
            if entity is None:
                return {}
            tags_list = entity.get("tags") or []
            return {t["key"]: t["values"] for t in tags_list}
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout, ConnectionResetError) as e:
            if attempt < 3:
                time.sleep(attempt * 5)
                continue
            raise
    return {}


# ---------------- NR: Add tags ----------------

def add_tags(api_key, nerdgraph_url, guid, tags):
    """Add tags to entity. tags = {key: value}. Returns list of errors."""
    tag_objs = [{"key": k, "values": [v]} for k, v in tags.items()]
    tags_gql = ", ".join(
        f'{{key: "{t["key"]}", values: {json.dumps(t["values"])}}}'
        for t in tag_objs
    )
    mutation = f"""
    mutation {{
      taggingAddTagsToEntity(guid: "{guid}", tags: [{tags_gql}]) {{
        errors {{ message type }}
      }}
    }}
    """
    for attempt in range(1, 4):
        try:
            response = requests.post(
                nerdgraph_url,
                json={"query": mutation},
                headers={"Api-Key": api_key, "Content-Type": "application/json"},
                timeout=30,
            )
            result = response.json()
            return (result.get("data") or {}).get("taggingAddTagsToEntity", {}).get("errors", []) or []
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout, ConnectionResetError) as e:
            if attempt < 3:
                time.sleep(attempt * 5)
                continue
            raise
    return []


# ---------------- NR: Delete tag keys ----------------

def delete_tag_keys(api_key, nerdgraph_url, guid, tag_keys):
    """Delete specific tag keys from an entity. Returns list of errors."""
    keys_gql = json.dumps(tag_keys)
    mutation = f"""
    mutation {{
      taggingDeleteTagFromEntity(guid: "{guid}", tagKeys: {keys_gql}) {{
        errors {{ message type }}
      }}
    }}
    """
    for attempt in range(1, 4):
        try:
            response = requests.post(
                nerdgraph_url,
                json={"query": mutation},
                headers={"Api-Key": api_key, "Content-Type": "application/json"},
                timeout=30,
            )
            result = response.json()
            return (result.get("data") or {}).get("taggingDeleteTagFromEntity", {}).get("errors", []) or []
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout, ConnectionResetError) as e:
            if attempt < 3:
                time.sleep(attempt * 5)
                continue
            raise
    return []


# ---------------- Comparison logic ----------------

def compare_tags(desired_tags, existing_tags):
    """Compare desired tags vs existing NR tags."""
    actions = {}
    for tag_key, desired_value in desired_tags.items():
        existing_values = existing_tags.get(tag_key, [])
        if not existing_values:
            actions[tag_key] = {"action": "add", "desired": desired_value, "existing": ""}
        elif len(existing_values) == 1 and existing_values[0] == desired_value:
            actions[tag_key] = {"action": "skip", "desired": desired_value, "existing": existing_values[0]}
        else:
            existing_str = ", ".join(existing_values)
            actions[tag_key] = {"action": "update", "desired": desired_value, "existing": existing_str}
    return actions


# ---------------- Push-only mode ----------------

def load_comparison_file(filepath):
    """Load comparison xlsx from a previous dry run.
    Returns list of dicts: [{guid, name, actions: {tag: {action, desired, existing}}}]
    """
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    results = []
    for row in rows[1:]:
        guid = row[0]
        name = row[1]
        actions = {}
        # Headers after guid, name are triplets: tag_action, tag_desired, tag_existing
        col = 2
        for tag_name in NR_TAG_NAMES:
            action_col = f"{tag_name}_action"
            if action_col in headers:
                idx = headers.index(action_col)
                action = row[idx] or ""
                desired = row[idx + 1] or ""
                existing = row[idx + 2] or ""
                if action in ("add", "update"):
                    actions[tag_name] = {"action": action, "desired": desired, "existing": existing}
        results.append({"guid": guid, "name": name, "actions": actions})
    wb.close()
    return results


def push_only_mode(config):
    """Push tags using a previously generated comparison file. Skips steps 1-5."""
    start_time = time.time()
    api_key = config["new_relic"]["api_key"]
    nerdgraph_url = config["new_relic"].get("nerdgraph_url", "https://api.newrelic.com/graphql")
    push_limit = config.get("push_limit", 0)
    push_skip = config.get("push_skip", 0)

    comp_file = config.get("push_only", "")
    if not os.path.isfile(comp_file):
        # Try as a path relative to extracts dir
        comp_file = os.path.join(EXTRACTS_DIR, comp_file)
    if not os.path.isfile(comp_file):
        raise FileNotFoundError(f"Comparison file not found: {config.get('push_only')}")

    print(f"=== Push-only mode ===")
    print(f"  Loading comparison: {comp_file}")
    entities = load_comparison_file(comp_file)
    print(f"  Loaded {len(entities)} entities")

    # Count actions
    total_to_push = sum(1 for e in entities if e["actions"])
    print(f"  Entities with tags to push: {total_to_push}")

    success_count = 0
    error_count = 0
    skipped = 0
    failed = []

    for i, ent in enumerate(entities, start=1):
        if push_skip and skipped < push_skip:
            skipped += 1
            continue

        if push_limit and success_count >= push_limit:
            print(f"  Push limit reached ({push_limit}). Stopping.")
            break

        actions = ent["actions"]
        if not actions:
            continue

        guid = ent["guid"]
        tags_to_add = {}
        tags_to_delete = []

        for tag_key, action_info in actions.items():
            if action_info["action"] == "add":
                tags_to_add[tag_key] = action_info["desired"]
            elif action_info["action"] == "update":
                tags_to_delete.append(tag_key)
                tags_to_add[tag_key] = action_info["desired"]

        if not tags_to_add and not tags_to_delete:
            continue

        try:
            if tags_to_delete:
                del_errs = delete_tag_keys(api_key, nerdgraph_url, guid, tags_to_delete)
                if del_errs:
                    print(f"  [{i}/{len(entities)}] DELETE errors for {guid}: {del_errs}")

            if tags_to_add:
                add_errs = add_tags(api_key, nerdgraph_url, guid, tags_to_add)
                if add_errs:
                    error_count += 1
                    failed.append({"guid": guid, "name": ent["name"], "errors": add_errs})
                    print(f"  [{i}/{len(entities)}] ADD errors: {add_errs}")
                else:
                    success_count += 1
                    if success_count % 20 == 0:
                        print(f"  [{i}/{len(entities)}] Tagged {success_count} so far...")
        except Exception as e:
            error_count += 1
            failed.append({"guid": guid, "name": ent["name"], "errors": [str(e)]})
            print(f"  [{i}/{len(entities)}] ERROR: {e}")

        if (success_count + error_count) % 50 == 0 and (success_count + error_count) > 0:
            print(f"  Batch of 50 processed. Sleeping 30s...")
            time.sleep(30)

    print(f"\n  Successfully tagged: {success_count}")
    print(f"  Errors: {error_count}")
    if failed:
        print("\n  Failed entities:")
        for entry in failed[:20]:
            print(f"    {entry['name']}: {entry['errors']}")

    elapsed = time.time() - start_time
    print(f"\n=== Done ({elapsed:.2f}s) ===")


# ---------------- Main ----------------

def main():
    start_time = time.time()
    config = load_config()

    # Push-only mode: skip steps 1-5, read from previous comparison file
    if config.get("push_only"):
        push_only_mode(config)
        return

    api_key = config["new_relic"]["api_key"]
    nerdgraph_url = config["new_relic"].get("nerdgraph_url", "https://api.newrelic.com/graphql")
    cmdb_base_url = config["cmdb_api"]["base_url"]
    cmdb_sub_key = config["cmdb_api"].get("subscription_key", "")
    apply = bool(config.get("apply_tags", False))

    os.makedirs(EXTRACTS_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # --- Step 1: Read input file ---
    print("=== Step 1: Read input file ===")
    if not os.path.isfile(_INPUT_FILE):
        raise FileNotFoundError(f"Input file not found: {_INPUT_FILE}")
    input_rows = parse_input_file(_INPUT_FILE)
    print(f"  Parsed {len(input_rows)} synthetics")
    unique_apps = set(r["business_app"] for r in input_rows)
    print(f"  Unique Business Applications: {len(unique_apps)}")

    # --- Step 2: Query CMDB API ---
    print(f"\n=== Step 2: Query CMDB API ({len(unique_apps)} apps) ===")
    cmdb_cache = {}
    for app_name in sorted(unique_apps):
        if not app_name.strip():
            print(f"  Skipping blank Business Application")
            continue
        print(f"  Querying: {app_name}")
        cmdb_data = query_cmdb(cmdb_base_url, app_name, cmdb_sub_key)
        if not cmdb_data and app_name in CMDB_ALIASES:
            alias = CMDB_ALIASES[app_name]
            print(f"    Retrying with alias: {alias}")
            cmdb_data = query_cmdb(cmdb_base_url, alias, cmdb_sub_key)
        if not cmdb_data and "|" in app_name:
            first_app = app_name.split("|")[0].strip()
            print(f"    Retrying with first value: {first_app}")
            cmdb_data = query_cmdb(cmdb_base_url, first_app, cmdb_sub_key)
        if cmdb_data:
            cmdb_cache[app_name] = cmdb_data
            print(f"    Found: {cmdb_data.get('mfc_app_code', '?')} - {cmdb_data.get('app_name', '?')}")
        else:
            print(f"    [WARN] No CMDB data found for '{app_name}'")
    print(f"  CMDB data retrieved for {len(cmdb_cache)}/{len(unique_apps)} apps")

    # --- Step 3: Look up synthetic GUIDs in NR ---
    print(f"\n=== Step 3: Look up GUIDs in NR ({len(input_rows)} synthetics) ===")
    entities = []
    not_found = []
    for i, row in enumerate(input_rows, start=1):
        name = row["name"]
        guid = search_synthetic_guid(api_key, nerdgraph_url, name)
        if guid:
            cmdb_data = cmdb_cache.get(row["business_app"])
            desired_tags = extract_tags_from_cmdb(cmdb_data, row["value_stream"])
            # Hardcode BusinessApplicationCIName for specific apps
            if row["business_app"] in HARDCODE_BACI_NAME:
                desired_tags["BusinessApplicationCIName"] = HARDCODE_BACI_NAME[row["business_app"]]
            entities.append({
                "name": name,
                "guid": guid,
                "business_app": row["business_app"],
                "desired_tags": desired_tags,
            })
        else:
            not_found.append(name)
            print(f"  [WARN] Not found in NR: {name}")
        if i % 20 == 0:
            print(f"  Looked up {i}/{len(input_rows)}...")

    print(f"  Found: {len(entities)}, Not found: {len(not_found)}")

    # --- Step 4: Query NR for existing tags + backup ---
    print(f"\n=== Step 4: Query NR for existing tags ({len(entities)} entities) ===")
    all_nr_tags = []
    for i, ent in enumerate(entities, start=1):
        nr_tags = query_entity_tags(api_key, nerdgraph_url, ent["guid"])
        all_nr_tags.append(nr_tags)
        if i % 20 == 0:
            print(f"  Queried {i}/{len(entities)}...")
    print(f"  Done querying {len(entities)} entities.")

    # Save backup
    backup_path = os.path.join(EXTRACTS_DIR, f"{timestamp}_gb_nr_tags_backup.xlsx")
    all_tag_keys = sorted(set(k for tags in all_nr_tags for k in tags))
    wb_backup = Workbook()
    ws_backup = wb_backup.active
    ws_backup.title = "NR Tags Backup"
    ws_backup.append(["nr_guid", "resource_name"] + all_tag_keys)
    for i, ent in enumerate(entities):
        row_data = [ent["guid"], ent["name"]]
        for key in all_tag_keys:
            vals = all_nr_tags[i].get(key, [])
            row_data.append(", ".join(vals) if vals else "")
        ws_backup.append(row_data)
    wb_backup.save(backup_path)
    print(f"  Backup saved: {backup_path}")

    # Push backup to SQL
    if config.get("push_to_sql", False):
        print("  Writing backup to SQL...")
        try:
            sql_config = config["sql"]
            pw = base64.b64decode(sql_config["pw_sql"]).decode("utf-8")
            conn_str = (
                f"DRIVER={{ODBC Driver 17 for SQL Server}};"
                f"SERVER={sql_config['server']};"
                f"DATABASE={sql_config['database']};"
                f"UID={sql_config['username_sql']};"
                f"PWD={pw};"
                f"Encrypt=yes;TrustServerCertificate=no;"
            )
            conn = pyodbc.connect(conn_str, timeout=30)
            cursor = conn.cursor()

            cursor.execute("""
                IF OBJECT_ID('dbo.tblSyntheticTagBackups_GB', 'U') IS NULL
                BEGIN
                    CREATE TABLE dbo.tblSyntheticTagBackups_GB (
                        recordID INT IDENTITY(1,1) PRIMARY KEY,
                        guid NVARCHAR(255),
                        tags NVARCHAR(MAX),
                        accountID NVARCHAR(50),
                        reportRunDate DATETIME,
                        recordSource NVARCHAR(50)
                    );
                END
            """)
            conn.commit()

            run_date = datetime.now()
            insert_sql = "INSERT INTO dbo.tblSyntheticTagBackups_GB (guid, tags, accountID, reportRunDate, recordSource) VALUES (?, ?, ?, ?, ?)"
            rows_inserted = 0
            for i, ent in enumerate(entities):
                guid = ent["guid"]
                nr_tags = all_nr_tags[i]
                tags_str = ",".join(
                    f"({k}:{', '.join(v)})" for k, v in nr_tags.items() if v
                )
                account_id = ", ".join(nr_tags.get("accountId", nr_tags.get("account", [])))
                cursor.execute(insert_sql, guid, tags_str, account_id, run_date, "SQL")
                rows_inserted += 1
            conn.commit()
            cursor.close()
            conn.close()
            print(f"  SQL backup complete: {rows_inserted} rows inserted into tblSyntheticTagBackups_GB")
        except Exception as e:
            print(f"  [WARN] SQL backup failed: {e}")
            print("  Continuing with xlsx backup only.")
    else:
        print("  SQL backup skipped (push_to_sql=false)")
        print("  Continuing with xlsx backup only.")

    # --- Step 5: Compare and generate outputs ---
    print("\n=== Step 5: Compare CMDB vs NR tags ===")
    total_adds = 0
    total_updates = 0
    total_skips = 0
    comparison_data = []

    for i, ent in enumerate(entities):
        desired = ent["desired_tags"]
        existing = all_nr_tags[i]
        actions = compare_tags(desired, existing)
        comparison_data.append(actions)
        for a in actions.values():
            if a["action"] == "add":
                total_adds += 1
            elif a["action"] == "update":
                total_updates += 1
            else:
                total_skips += 1

    print(f"  Actions summary: {total_adds} adds, {total_updates} updates, {total_skips} skips")

    # Write comparison file
    comparison_path = os.path.join(EXTRACTS_DIR, f"{timestamp}_gb_tag_comparison.xlsx")
    wb_comp = Workbook()
    ws_comp = wb_comp.active
    ws_comp.title = "Tag Comparison"
    comp_headers = ["nr_guid", "resource_name"]
    for tag_name in NR_TAG_NAMES:
        comp_headers.extend([f"{tag_name}_action", f"{tag_name}_desired", f"{tag_name}_existing"])
    ws_comp.append(comp_headers)
    for i, ent in enumerate(entities):
        row_data = [ent["guid"], ent["name"]]
        actions = comparison_data[i]
        existing = all_nr_tags[i]
        for tag_name in NR_TAG_NAMES:
            if tag_name in actions:
                a = actions[tag_name]
                row_data.extend([a["action"], a["desired"], a["existing"]])
            else:
                existing_vals = existing.get(tag_name, [])
                row_data.extend(["no_change (source empty)", "", ", ".join(existing_vals) if existing_vals else ""])
        ws_comp.append(row_data)
    wb_comp.save(comparison_path)
    print(f"  Comparison: {comparison_path}")

    # Write push preview
    push_preview_path = os.path.join(EXTRACTS_DIR, f"{timestamp}_gb_push_preview.xlsx")
    wb_push = Workbook()
    ws_push = wb_push.active
    ws_push.title = "Push Preview"
    ws_push.append(["nr_guid", "resource_name"] + NR_TAG_NAMES)
    for i, ent in enumerate(entities):
        row_data = [ent["guid"], ent["name"]]
        actions = comparison_data[i]
        for tag_name in NR_TAG_NAMES:
            if tag_name in actions and actions[tag_name]["action"] in ("add", "update"):
                row_data.append(actions[tag_name]["desired"])
            else:
                row_data.append("")
        ws_push.append(row_data)
    wb_push.save(push_preview_path)
    print(f"  Push preview: {push_preview_path}")

    # Log not-found synthetics
    if not_found:
        not_found_path = os.path.join(EXTRACTS_DIR, f"{timestamp}_gb_not_found.txt")
        with open(not_found_path, "w") as f:
            for name in not_found:
                f.write(name + "\n")
        print(f"  Not found list: {not_found_path}")

    # --- Step 6: Push changes ---
    print(f"\n=== Step 6: Push tags (apply_tags={apply}) ===")
    if not apply:
        print("  Dry run — no tags will be pushed. Review the output files.")
        print("  Set apply_tags=true in config.json to push.")
    else:
        push_limit = config.get("push_limit", 0)
        push_skip = config.get("push_skip", 0)
        success_count = 0
        error_count = 0
        skipped = 0
        failed = []

        for i, ent in enumerate(entities, start=1):
            if push_skip and skipped < push_skip:
                skipped += 1
                continue

            if push_limit and success_count >= push_limit:
                print(f"  Push limit reached ({push_limit}). Stopping.")
                break

            guid = ent["guid"]
            actions = comparison_data[i - 1]

            tags_to_add = {}
            tags_to_delete = []

            for tag_key, action_info in actions.items():
                if action_info["action"] == "add":
                    tags_to_add[tag_key] = action_info["desired"]
                elif action_info["action"] == "update":
                    tags_to_delete.append(tag_key)
                    tags_to_add[tag_key] = action_info["desired"]

            if not tags_to_add and not tags_to_delete:
                continue

            try:
                if tags_to_delete:
                    del_errs = delete_tag_keys(api_key, nerdgraph_url, guid, tags_to_delete)
                    if del_errs:
                        print(f"  [{i}/{len(entities)}] DELETE errors for {guid}: {del_errs}")

                if tags_to_add:
                    add_errs = add_tags(api_key, nerdgraph_url, guid, tags_to_add)
                    if add_errs:
                        error_count += 1
                        failed.append({"guid": guid, "name": ent["name"], "errors": add_errs})
                        print(f"  [{i}/{len(entities)}] ADD errors: {add_errs}")
                    else:
                        success_count += 1
                        if success_count % 20 == 0:
                            print(f"  [{i}/{len(entities)}] Tagged {success_count} so far...")
            except Exception as e:
                error_count += 1
                failed.append({"guid": guid, "name": ent["name"], "errors": [str(e)]})
                print(f"  [{i}/{len(entities)}] ERROR: {e}")

            # Rate limiting
            if (success_count + error_count) % 50 == 0 and (success_count + error_count) > 0:
                print(f"  Batch of 50 processed. Sleeping 30s...")
                time.sleep(30)

        print(f"\n  Successfully tagged: {success_count}")
        print(f"  Errors: {error_count}")
        if failed:
            print("\n  Failed entities:")
            for entry in failed[:20]:
                print(f"    {entry['name']}: {entry['errors']}")

    elapsed = time.time() - start_time
    print(f"\n=== Done ({elapsed:.2f}s) ===")
    print(f"  Backup: {backup_path}")
    print(f"  Comparison: {comparison_path}")
    print(f"  Push preview: {push_preview_path}")


if __name__ == "__main__":
    main()
