"""MMM Entity tagger (v3).

Flow:
  1. Read input CSV (nr_guid + 7 tag columns)
  2. Query NR for existing tags → save backup
  3. Compare desired vs existing → generate comparison + push preview
  4. Push changes when apply_tags=true

Configuration: ./config/config.json
"""

import csv
import json
import os
import time
from datetime import datetime

import requests
from openpyxl import Workbook

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_CONFIG_PATH = os.path.join(_SCRIPT_DIR, "config", "config.json")
_INPUT_FILE = os.path.join(_SCRIPT_DIR, "MMM Tagging(GB).csv")
EXTRACTS_DIR = os.path.join(_SCRIPT_DIR, "extracts")

# Tag keys from the CSV (mapped to NR tag names)
TAG_COLUMNS = [
    "Segment",
    "BU",
    "MFCAppCode",
    "MFCAppService",
    "critical_classification",
    "Monitoring",
    "CriticalityClassification",
    "DigitalPerformance",
]


def load_config():
    with open(_CONFIG_PATH) as f:
        return json.load(f)


# ---------------- Parse input CSV ----------------

def parse_input_file(filepath):
    """Read CSV and return list of dicts with guid + tags."""
    rows = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            guid = row.get("entity_guid", "").strip()
            if not guid:
                continue
            tags = {}
            for col in TAG_COLUMNS:
                val = row.get(col, "").strip()
                if val:
                    tags[col] = val
            rows.append({"guid": guid, "tags": tags})
    return rows


# ---------------- NR: Query existing tags ----------------

def query_entity_tags(api_key, nerdgraph_url, guid):
    """Query NR for all tags on an entity. Returns dict {tag_key: [values]}."""
    query = """
    {
      actor {
        entity(guid: "%s") {
          tags {
            key
            values
          }
        }
      }
    }
    """ % guid

    for attempt in range(1, 4):
        try:
            response = requests.post(
                nerdgraph_url,
                json={"query": query},
                headers={"Api-Key": api_key, "Content-Type": "application/json"},
                timeout=30,
            )
            result = response.json()
            entity = (result.get("data") or {}).get("actor", {}).get("entity")
            if entity is None:
                return {}
            tags_list = entity.get("tags") or []
            return {t["key"]: t["values"] for t in tags_list}
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout, ConnectionResetError) as e:
            if attempt < 3:
                time.sleep(attempt * 5)
                continue
            raise
    return {}


# ---------------- NR: Add tags ----------------

def add_tags(api_key, nerdgraph_url, guid, tags):
    """Add tags to entity. tags = {key: value}. Returns list of errors."""
    tag_objs = [{"key": k, "values": [v]} for k, v in tags.items()]
    tags_gql = ", ".join(
        f'{{key: "{t["key"]}", values: {json.dumps(t["values"])}}}'
        for t in tag_objs
    )
    mutation = f"""
    mutation {{
      taggingAddTagsToEntity(guid: "{guid}", tags: [{tags_gql}]) {{
        errors {{ message type }}
      }}
    }}
    """
    for attempt in range(1, 4):
        try:
            response = requests.post(
                nerdgraph_url,
                json={"query": mutation},
                headers={"Api-Key": api_key, "Content-Type": "application/json"},
                timeout=30,
            )
            result = response.json()
            return (result.get("data") or {}).get("taggingAddTagsToEntity", {}).get("errors", []) or []
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout, ConnectionResetError) as e:
            if attempt < 3:
                time.sleep(attempt * 5)
                continue
            raise
    return []


# ---------------- NR: Delete tag keys ----------------

def delete_tag_keys(api_key, nerdgraph_url, guid, tag_keys):
    """Delete specific tag keys from an entity. Returns list of errors."""
    keys_gql = json.dumps(tag_keys)
    mutation = f"""
    mutation {{
      taggingDeleteTagFromEntity(guid: "{guid}", tagKeys: {keys_gql}) {{
        errors {{ message type }}
      }}
    }}
    """
    for attempt in range(1, 4):
        try:
            response = requests.post(
                nerdgraph_url,
                json={"query": mutation},
                headers={"Api-Key": api_key, "Content-Type": "application/json"},
                timeout=30,
            )
            result = response.json()
            return (result.get("data") or {}).get("taggingDeleteTagFromEntity", {}).get("errors", []) or []
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout, ConnectionResetError) as e:
            if attempt < 3:
                time.sleep(attempt * 5)
                continue
            raise
    return []


# ---------------- Comparison logic ----------------

def compare_tags(desired_tags, existing_tags):
    """Compare desired tags vs existing NR tags.

    Returns dict: {tag_key: {'action': 'add'|'update'|'skip', 'desired': str, 'existing': str}}
    """
    actions = {}
    for tag_key, desired_value in desired_tags.items():
        existing_values = existing_tags.get(tag_key, [])
        if not existing_values:
            actions[tag_key] = {"action": "add", "desired": desired_value, "existing": ""}
        elif len(existing_values) == 1 and existing_values[0] == desired_value:
            actions[tag_key] = {"action": "skip", "desired": desired_value, "existing": existing_values[0]}
        else:
            existing_str = ", ".join(existing_values)
            actions[tag_key] = {"action": "update", "desired": desired_value, "existing": existing_str}
    return actions


# ---------------- Main ----------------

def main():
    start_time = time.time()
    config = load_config()
    api_key = config["new_relic"]["api_key"]
    nerdgraph_url = config["new_relic"].get("nerdgraph_url", "https://api.newrelic.com/graphql")
    apply = bool(config.get("apply_tags", False))
    push_limit = config.get("push_limit", 0)
    push_skip = config.get("push_skip", 0)

    os.makedirs(EXTRACTS_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # --- Step 1: Read input file ---
    print("=== Step 1: Read input file ===")
    if not os.path.isfile(_INPUT_FILE):
        raise FileNotFoundError(f"Input file not found: {_INPUT_FILE}")
    input_rows = parse_input_file(_INPUT_FILE)
    print(f"  Parsed {len(input_rows)} entities with GUIDs")

    # --- Step 2: Query NR for existing tags ---
    print(f"\n=== Step 2: Query NR for existing tags ({len(input_rows)} entities) ===")
    all_nr_tags = []
    for i, row in enumerate(input_rows, start=1):
        existing = query_entity_tags(api_key, nerdgraph_url, row["guid"])
        all_nr_tags.append(existing)
        if i % 20 == 0:
            print(f"  Queried {i}/{len(input_rows)}...")
    print(f"  Done querying {len(input_rows)} entities.")

    # Save backup xlsx
    backup_path = os.path.join(EXTRACTS_DIR, f"{timestamp}_mmm_tags_backup.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "NR Tags Backup"
    ws.append(["guid", "tags"])
    for i, row in enumerate(input_rows):
        nr_tags = all_nr_tags[i]
        tags_str = ",".join(f"({k}:{', '.join(v)})" for k, v in nr_tags.items() if v)
        ws.append([row["guid"], tags_str])
    wb.save(backup_path)
    print(f"  Backup saved: {backup_path}")

    # --- Step 3: Compare desired vs existing ---
    print(f"\n=== Step 3: Compare desired vs existing tags ===")
    comparison_data = []
    total_adds = 0
    total_updates = 0
    total_skips = 0

    for i, row in enumerate(input_rows):
        existing = all_nr_tags[i]
        actions = compare_tags(row["tags"], existing)
        comparison_data.append(actions)
        for a in actions.values():
            if a["action"] == "add":
                total_adds += 1
            elif a["action"] == "update":
                total_updates += 1
            else:
                total_skips += 1

    print(f"  Actions summary: {total_adds} adds, {total_updates} updates, {total_skips} skips")

    # Save comparison xlsx
    comp_path = os.path.join(EXTRACTS_DIR, f"{timestamp}_mmm_tag_comparison.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    ws.append(["guid", "tag_key", "action", "desired_value", "existing_value"])
    for i, row in enumerate(input_rows):
        for tag_key, info in comparison_data[i].items():
            ws.append([row["guid"], tag_key, info["action"], info["desired"], info["existing"]])
    wb.save(comp_path)
    print(f"  Comparison: {comp_path}")

    # Save push preview xlsx (all tags being pushed)
    preview_path = os.path.join(EXTRACTS_DIR, f"{timestamp}_mmm_push_preview.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Push Preview"
    ws.append(["guid", "tag_key", "action", "value_to_push"])
    for i, row in enumerate(input_rows):
        for tag_key, desired_value in row["tags"].items():
            ws.append([row["guid"], tag_key, "add", desired_value])
    wb.save(preview_path)
    print(f"  Push preview: {preview_path}")

    # --- Step 4: Push tags ---
    print(f"\n=== Step 4: Push tags (apply_tags={apply}) ===")
    if not apply:
        print("  Dry run — no tags will be pushed. Review the output files.")
        print("  Set apply_tags=true in config.json to push.")
    else:
        success_count = 0
        error_count = 0
        failed = []

        for i, row in enumerate(input_rows, start=1):
            if push_skip and i <= push_skip:
                continue
            if push_limit and success_count >= push_limit:
                print(f"  Push limit reached ({push_limit}). Stopping.")
                break

            guid = row["guid"]
            actions = comparison_data[i - 1]

            # Push ALL tags unconditionally (add only, no deletes)
            tags_to_add = row["tags"]

            if not tags_to_add:
                continue

            try:
                if tags_to_add:
                    add_errs = add_tags(api_key, nerdgraph_url, guid, tags_to_add)
                    if add_errs:
                        error_count += 1
                        failed.append({"guid": guid, "errors": add_errs})
                        print(f"  [{i}/{len(input_rows)}] ADD errors: {add_errs}")
                    else:
                        success_count += 1
                        if success_count % 20 == 0:
                            print(f"  [{i}/{len(input_rows)}] Tagged {success_count} so far...")
            except Exception as e:
                error_count += 1
                failed.append({"guid": guid, "errors": [str(e)]})
                print(f"  [{i}/{len(input_rows)}] ERROR: {e}")

            # Rate limit: sleep after every 50 pushes
            if success_count > 0 and success_count % 50 == 0:
                print(f"  Batch of 50 processed. Sleeping 30s...")
                time.sleep(30)

        print(f"\n  Successfully tagged: {success_count}")
        print(f"  Errors: {error_count}")
        if failed:
            failed_path = os.path.join(EXTRACTS_DIR, f"{timestamp}_mmm_failed.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "Failed"
            ws.append(["guid", "errors"])
            for f_item in failed:
                ws.append([f_item["guid"], str(f_item["errors"])])
            wb.save(failed_path)
            print(f"  Failed details: {failed_path}")

    elapsed = time.time() - start_time
    print(f"\n=== Done ({elapsed:.2f}s) ===")
    print(f"  Backup: {backup_path}")
    print(f"  Comparison: {comp_path}")
    print(f"  Push preview: {preview_path}")


if __name__ == "__main__":
    main()
