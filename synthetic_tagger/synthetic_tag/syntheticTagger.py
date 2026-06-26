"""Synthetic entity tagger.

Reads synthetics_export.xlsx (source of truth), queries New Relic for existing
tags on each entity, compares, and reconciles differences.

Flow:
  1. Parse input Excel
  2. Query NR for existing tags → save backup xlsx
  3. Compare export vs NR → generate comparison preview xlsx
  4. Push changes (add/delete+add) only when apply_tags=true

Configuration: ./config/config.json
"""

import json
import os
import time
from datetime import datetime

import requests
import openpyxl
from openpyxl import Workbook

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_CONFIG_PATH = os.path.join(_SCRIPT_DIR, "config", "config.json")
_INPUT_FILE = os.path.join(_SCRIPT_DIR, "synthetics_export.xlsx")
EXTRACTS_DIR = os.path.join(_SCRIPT_DIR, "extracts")

# Excel column → NR tag name mapping
TAG_MAPPING = {
    "app_code": ["BusinessApplicationCode", "MFCAppCode"],
    "app_name": ["BusinessApplicationCIName"],
    "app_ci_number": ["BusinessApplicationCINumber", "MFCAppService"],
    "app_service_name": ["ApplicationServiceCIName"],
    "app_service_ci": ["ApplicationServiceCINumber"],
    "leanix_segment": ["Segment"],
    "leanix_bu": ["Bu"],
    "application_criticality": ["critical_classification"],
}

GUID_COLUMN_CANDIDATES = ("guid", "nr_guid", "entity_guid", "entityguid")
NAME_COLUMN_CANDIDATES = ("name", "resource_name", "monitor_name")


def load_config():
    with open(_CONFIG_PATH) as f:
        return json.load(f)


def clean_criticality(value):
    """Strip '2-', '3-', '4-' prefix from criticality values."""
    if not value:
        return value
    for prefix in ("2-", "3-", "4-"):
        if value.startswith(prefix):
            return value[len(prefix):]
    return value


# ---------------- Excel parsing ----------------

def parse_input_file(filepath):
    """Parse the input Excel. Returns list of dicts with guid, name, and desired tags.

    Uses TAG_MAPPING to convert excel columns to NR tag names.
    Deduplicates by GUID (first occurrence wins).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    # Find GUID column
    guid_idx = None
    for i, h in enumerate(headers):
        if h.lower() in GUID_COLUMN_CANDIDATES:
            guid_idx = i
            break
    if guid_idx is None:
        raise RuntimeError(f"No GUID column found. Headers: {headers}")

    # Find name column
    name_idx = None
    for i, h in enumerate(headers):
        if h.lower() in NAME_COLUMN_CANDIDATES:
            name_idx = i
            break

    # Map excel columns to their indices
    col_indices = {}
    for col_name in TAG_MAPPING:
        for i, h in enumerate(headers):
            if h == col_name:
                col_indices[col_name] = i
                break

    seen_guids = set()
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or guid_idx >= len(row):
            continue
        guid = row[guid_idx]
        if not guid:
            continue
        guid = str(guid).strip()

        # Deduplicate: first occurrence wins
        if guid in seen_guids:
            continue
        seen_guids.add(guid)

        name = ""
        if name_idx is not None and name_idx < len(row) and row[name_idx]:
            name = str(row[name_idx]).strip()

        # Build desired tags using TAG_MAPPING
        desired_tags = {}
        for excel_col, nr_tag_names in TAG_MAPPING.items():
            idx = col_indices.get(excel_col)
            if idx is None or idx >= len(row) or row[idx] is None:
                continue
            val = str(row[idx]).strip()
            if not val:
                continue
            for nr_tag in nr_tag_names:
                if nr_tag == "critical_classification":
                    desired_tags[nr_tag] = clean_criticality(val)
                else:
                    desired_tags[nr_tag] = val

        rows.append({"guid": guid, "name": name, "desired_tags": desired_tags})
    wb.close()
    return rows


# ---------------- NR: Query existing tags ----------------

def query_entity_tags(api_key, nerdgraph_url, guid):
    """Query NR for all tags on an entity. Returns dict {tag_key: [values]}."""
    query = """
    {
      actor {
        entity(guid: "%s") {
          tags {
            key
            values
          }
        }
      }
    }
    """ % guid

    for attempt in range(1, 4):
        try:
            response = requests.post(
                nerdgraph_url,
                json={"query": query},
                headers={"Api-Key": api_key, "Content-Type": "application/json"},
                timeout=30,
            )
            result = response.json()
            entity = (result.get("data") or {}).get("actor", {}).get("entity")
            if entity is None:
                return {}
            tags_list = entity.get("tags") or []
            return {t["key"]: t["values"] for t in tags_list}
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout, ConnectionResetError) as e:
            if attempt < 3:
                time.sleep(attempt * 5)
                continue
            raise
    return {}


# ---------------- NR: Add tags ----------------

def add_tags(api_key, nerdgraph_url, guid, tags):
    """Add tags to entity. tags = {key: value}. Returns list of errors."""
    tag_objs = [{"key": k, "values": [v]} for k, v in tags.items()]
    tags_gql = ", ".join(
        f'{{key: "{t["key"]}", values: {json.dumps(t["values"])}}}'
        for t in tag_objs
    )
    mutation = f"""
    mutation {{
      taggingAddTagsToEntity(guid: "{guid}", tags: [{tags_gql}]) {{
        errors {{ message type }}
      }}
    }}
    """
    for attempt in range(1, 4):
        response = None
        try:
            response = requests.post(
                nerdgraph_url,
                json={"query": mutation},
                headers={"Api-Key": api_key, "Content-Type": "application/json"},
                timeout=30,
            )
            result = response.json()
            return (result.get("data") or {}).get("taggingAddTagsToEntity", {}).get("errors", []) or []
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout, ConnectionResetError) as e:
            if attempt < 3:
                time.sleep(attempt * 5)
                continue
            raise
    return []


# ---------------- NR: Delete tag values ----------------

def delete_tag_values(api_key, nerdgraph_url, guid, tag_keys):
    """Delete specific tag keys from an entity. Returns list of errors."""
    keys_gql = json.dumps(tag_keys)
    mutation = f"""
    mutation {{
      taggingDeleteTagFromEntity(guid: "{guid}", tagKeys: {keys_gql}) {{
        errors {{ message type }}
      }}
    }}
    """
    for attempt in range(1, 4):
        response = None
        try:
            response = requests.post(
                nerdgraph_url,
                json={"query": mutation},
                headers={"Api-Key": api_key, "Content-Type": "application/json"},
                timeout=30,
            )
            result = response.json()
            return (result.get("data") or {}).get("taggingDeleteTagFromEntity", {}).get("errors", []) or []
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout, ConnectionResetError) as e:
            if attempt < 3:
                time.sleep(attempt * 5)
                continue
            raise
    return []


# ---------------- Comparison logic ----------------

def compare_tags(desired_tags, existing_tags):
    """Compare desired tags vs existing NR tags.

    Returns dict of actions: {tag_key: {'action': 'add'|'update'|'skip', 'desired': str, 'existing': str}}
    - 'add': tag doesn't exist in NR
    - 'update': tag exists but value differs (will delete then add)
    - 'skip': tag already matches
    Only considers tags present in desired_tags (empty export values already excluded).
    """
    actions = {}
    for tag_key, desired_value in desired_tags.items():
        existing_values = existing_tags.get(tag_key, [])
        # NR returns values as a list; check if desired matches any single value
        if not existing_values:
            actions[tag_key] = {"action": "add", "desired": desired_value, "existing": ""}
        elif len(existing_values) == 1 and existing_values[0] == desired_value:
            actions[tag_key] = {"action": "skip", "desired": desired_value, "existing": existing_values[0]}
        else:
            # Value differs or multi-value — need to replace
            existing_str = ", ".join(existing_values)
            actions[tag_key] = {"action": "update", "desired": desired_value, "existing": existing_str}
    return actions


# ---------------- Main ----------------

def main():
    start_time = time.time()
    config = load_config()
    api_key = config["new_relic"]["api_key"]
    nerdgraph_url = config["new_relic"].get("nerdgraph_url", "https://api.newrelic.com/graphql")
    apply = bool(config.get("apply_tags", False))

    # --- Step 1: Read input file ---
    print("=== Step 1: Read synthetics_export.xlsx ===")
    if not os.path.isfile(_INPUT_FILE):
        raise FileNotFoundError(f"Expected input file not found: {_INPUT_FILE}")
    rows = parse_input_file(_INPUT_FILE)
    print(f"  Parsed {len(rows)} unique entities")
    print(f"  Target NR tags: {list(set(t for tags in TAG_MAPPING.values() for t in tags))}")

    # --- Step 2: Query NR for existing tags + backup ---
    print(f"\n=== Step 2: Query NR for existing tags ({len(rows)} entities) ===")
    os.makedirs(EXTRACTS_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    all_nr_tags = []  # list of {guid, name, nr_tags_dict}
    for i, row in enumerate(rows, start=1):
        guid = row["guid"]
        nr_tags = query_entity_tags(api_key, nerdgraph_url, guid)
        all_nr_tags.append({"guid": guid, "name": row["name"], "nr_tags": nr_tags})
        if i % 50 == 0:
            print(f"  Queried {i}/{len(rows)}...")

    print(f"  Done querying {len(rows)} entities.")

    # Save backup of existing NR tags
    backup_path = os.path.join(EXTRACTS_DIR, f"{timestamp}_nr_tags_backup.xlsx")
    all_tag_keys = sorted(set(k for entry in all_nr_tags for k in entry["nr_tags"]))
    wb_backup = Workbook()
    ws_backup = wb_backup.active
    ws_backup.title = "NR Tags Backup"
    ws_backup.append(["nr_guid", "resource_name"] + all_tag_keys)
    for entry in all_nr_tags:
        row_data = [entry["guid"], entry["name"]]
        for key in all_tag_keys:
            vals = entry["nr_tags"].get(key, [])
            row_data.append(", ".join(vals) if vals else "")
        ws_backup.append(row_data)
    wb_backup.save(backup_path)
    print(f"  Backup saved: {backup_path}")

    # --- Step 3: Compare and generate preview ---
    print("\n=== Step 3: Compare export vs NR tags ===")
    nr_tag_names = sorted(set(t for tags in TAG_MAPPING.values() for t in tags))

    comparison_rows = []  # For preview output
    total_adds = 0
    total_updates = 0
    total_skips = 0

    for i, row in enumerate(rows):
        guid = row["guid"]
        name = row["name"]
        desired = row["desired_tags"]
        existing = all_nr_tags[i]["nr_tags"]
        actions = compare_tags(desired, existing)

        comp_row = {"guid": guid, "name": name}
        for tag_name in nr_tag_names:
            if tag_name in actions:
                a = actions[tag_name]
                comp_row[f"{tag_name}_action"] = a["action"]
                comp_row[f"{tag_name}_desired"] = a["desired"]
                comp_row[f"{tag_name}_existing"] = a["existing"]
                if a["action"] == "add":
                    total_adds += 1
                elif a["action"] == "update":
                    total_updates += 1
                else:
                    total_skips += 1
            else:
                # Export value was empty — not touching this tag
                existing_vals = existing.get(tag_name, [])
                comp_row[f"{tag_name}_action"] = "no_change (export empty)"
                comp_row[f"{tag_name}_desired"] = ""
                comp_row[f"{tag_name}_existing"] = ", ".join(existing_vals) if existing_vals else ""

        comparison_rows.append(comp_row)

    # Write comparison preview
    preview_path = os.path.join(EXTRACTS_DIR, f"{timestamp}_tag_comparison.xlsx")
    wb_preview = Workbook()
    ws_preview = wb_preview.active
    ws_preview.title = "Tag Comparison"

    # Build headers: nr_guid, resource_name, then for each tag: tag_action, tag_desired, tag_existing
    preview_headers = ["nr_guid", "resource_name"]
    for tag_name in nr_tag_names:
        preview_headers.extend([f"{tag_name}_action", f"{tag_name}_desired", f"{tag_name}_existing"])
    ws_preview.append(preview_headers)

    for comp in comparison_rows:
        row_data = [comp["guid"], comp["name"]]
        for tag_name in nr_tag_names:
            row_data.append(comp.get(f"{tag_name}_action", ""))
            row_data.append(comp.get(f"{tag_name}_desired", ""))
            row_data.append(comp.get(f"{tag_name}_existing", ""))
        ws_preview.append(row_data)
    wb_preview.save(preview_path)

    print(f"  Comparison preview: {preview_path}")
    print(f"  Actions summary: {total_adds} adds, {total_updates} updates, {total_skips} skips")

    # Write clean push preview (only shows values that will be pushed, empty if already correct)
    push_preview_path = os.path.join(EXTRACTS_DIR, f"{timestamp}_push_preview.xlsx")
    wb_push = Workbook()
    ws_push = wb_push.active
    ws_push.title = "Push Preview"
    ws_push.append(["nr_guid", "resource_name"] + nr_tag_names)
    for i, row in enumerate(rows):
        guid = row["guid"]
        name = row["name"]
        desired = row["desired_tags"]
        existing = all_nr_tags[i]["nr_tags"]
        actions = compare_tags(desired, existing)
        row_data = [guid, name]
        for tag_name in nr_tag_names:
            if tag_name in actions and actions[tag_name]["action"] in ("add", "update"):
                row_data.append(actions[tag_name]["desired"])
            else:
                row_data.append("")
        ws_push.append(row_data)
    wb_push.save(push_preview_path)
    print(f"  Push preview: {push_preview_path}")

    # --- Step 4: Push changes (only if apply_tags=true) ---
    print(f"\n=== Step 4: Push tags (apply_tags={apply}) ===")
    if not apply:
        print("  Dry run — no tags will be pushed. Review the comparison file.")
        print("  Set apply_tags=true in config.json to push.")
    else:
        success_count = 0
        error_count = 0
        failed = []

        for i, row in enumerate(rows, start=1):
            guid = row["guid"]
            desired = row["desired_tags"]
            existing = all_nr_tags[i - 1]["nr_tags"]
            actions = compare_tags(desired, existing)

            tags_to_add = {}
            tags_to_delete = []

            for tag_key, action_info in actions.items():
                if action_info["action"] == "add":
                    tags_to_add[tag_key] = action_info["desired"]
                elif action_info["action"] == "update":
                    tags_to_delete.append(tag_key)
                    tags_to_add[tag_key] = action_info["desired"]
                # skip → do nothing

            if not tags_to_add and not tags_to_delete:
                continue

            try:
                # Delete tags that need updating
                if tags_to_delete:
                    del_errs = delete_tag_values(api_key, nerdgraph_url, guid, tags_to_delete)
                    if del_errs:
                        print(f"  [{i}/{len(rows)}] DELETE errors for {guid}: {del_errs}")

                # Add new/updated tags
                if tags_to_add:
                    add_errs = add_tags(api_key, nerdgraph_url, guid, tags_to_add)
                    if add_errs:
                        error_count += 1
                        failed.append({"guid": guid, "errors": add_errs})
                        print(f"  [{i}/{len(rows)}] ADD errors for {guid}: {add_errs}")
                    else:
                        success_count += 1
                        if success_count % 50 == 0:
                            print(f"  [{i}/{len(rows)}] Tagged {success_count} so far...")

            except Exception as e:
                error_count += 1
                failed.append({"guid": guid, "errors": [str(e)]})
                print(f"  [{i}/{len(rows)}] ERROR for {guid}: {e}")

            # Rate limiting: pause every 100 operations
            if (success_count + error_count) % 100 == 0 and (success_count + error_count) > 0:
                print(f"  Batch of 100 processed. Sleeping 30s...")
                time.sleep(30)

        print(f"\n  Successfully tagged: {success_count}")
        print(f"  Errors: {error_count}")
        if failed:
            print("\n  Failed entities:")
            for entry in failed[:20]:
                print(f"    {entry['guid']}: {entry['errors']}")

    elapsed = time.time() - start_time
    print(f"\n=== Done ({elapsed:.2f}s) ===")
    print(f"  Backup: {backup_path}")
    print(f"  Comparison: {preview_path}")
    print(f"  Push preview: {push_preview_path}")


if __name__ == "__main__":
    main()
